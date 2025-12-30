import json
import os
import logging
import asyncio
from typing import Dict, Any
import re

from app.pdf_locator import find_text_coordinates
from google import genai
import openai
from PyPDF2 import PdfReader
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Configure logging
# These logs will be useful for a future UI-based developer log window
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constants
DEFAULT_GEMINI_MODEL = "gemini-2.5-flash"
DEFAULT_OPENAI_MODEL = "gpt-4o"

# Mock data for "Test Mode" to avoid API costs during UI development.
# These examples represent typical legal issues found in commercial contracts.
MOCK_ANALYSIS_RESULT = {
  "errors": [
    {
      "location": "Page 3, Section 2.1(a)",
      "error": "The capitalized term \"Amendment\" is used but not defined. The document defines \"this Agreement\" (capitalized) to refer to the First Amendment to Amended and Restated Credit Agreement, so \"Amendment\" should likely be \"Agreement\" for consistency.",
      "suggestion": "Replace \"this Amendment\" with \"this Agreement\".",
      "exact_quote": "delivery and performance of this Amendment",
      "boundingBoxes": [{"x": 72, "y": 150, "width": 200, "height": 12, "page": 3, "page_width": 595, "page_height": 842}]
    },
    {
      "location": "Page 3, Section 3(a)(ii)",
      "error": "The sentence ends abruptly and carries over to the next page without proper continuation, indicating a drafting error or forgotten placeholder.",
      "suggestion": "Complete the sentence or ensure proper pagination and sentence flow.",
      "exact_quote": "certificates as of a recent date of the good standing of each Credit Party",
      "boundingBoxes": [{"x": 72, "y": 700, "width": 300, "height": 12, "page": 3, "page_width": 595, "page_height": 842}]
    },
    {
      "location": "Page 4, Section 3(a)(iii)",
      "error": "The sentence ends abruptly and carries over to the next page without proper continuation, indicating a drafting error or forgotten placeholder.",
      "suggestion": "Complete the sentence or ensure proper pagination and sentence flow.",
      "exact_quote": "the Borrower is in compliance with the",
      "boundingBoxes": [{"x": 72, "y": 700, "width": 250, "height": 12, "page": 4, "page_width": 595, "page_height": 842}]
    },
    {
      "location": "Page 5, Section 3(b)",
      "error": "The condition refers to 'Section 3' for representations and warranties, but the representations and warranties are actually set forth in 'Section 4'. This appears to be a numbering error.",
      "suggestion": "Change 'Section 3' to 'Section 4'.",
      "exact_quote": "The representations and warranties set forth in Section 3 shall be true and correct."
    },
    {
      "location": "Page 9",
      "error": "A placeholder indicates that signature pages are expected to follow, which is a drafting error in a completed document.",
      "suggestion": "Remove the placeholder or replace it with appropriate text if the document is finalized.",
      "exact_quote": "[Signature pages to follow]"
    },
    {
      "location": "Page 47, Article I (Definitions)",
      "error": "There is a pagination error. The page number is '13', but it should sequentially be '14' after the preceding page (Page 46) which was '13'.",
      "suggestion": "Correct the page number to '14'.",
      "exact_quote": "13"
    },
    {
      "location": "Page 49, Article I (Definitions)",
      "error": "There is a pagination error. The page number is '14', but it should sequentially be '15' after the preceding page (Page 48) which was '14'.",
      "suggestion": "Correct the page number to '15'.",
      "exact_quote": "14"
    },
    {
      "location": "Page 54, Article I (Definitions)",
      "error": "The singular term 'Note' is capitalized and used in the definition of 'Loan Documents' but is not explicitly defined. While 'Notes' (plural) is defined later, consistency requires either 'Note' to be defined or 'Notes' to be used consistently.",
      "suggestion": "Define 'Note' as the singular of 'Notes' or replace 'each Note' with 'the Notes' in the definition of 'Loan Documents'.",
      "exact_quote": "each Note"
    },
    {
      "location": "Page 55, Article I (Definitions)",
      "error": "The capitalized term 'Outstanding Amount' is used in the definition of 'Minimum Collateral Amount' but is not defined.",
      "suggestion": "Add a definition for 'Outstanding Amount'.",
      "exact_quote": "Outstanding Amount"
    },
    {
      "location": "Page 60, Article I (Definitions)",
      "error": "There is a typographical error in 'Sanctioned Peron(s)'. It should be 'Sanctioned Person(s)'.",
      "suggestion": "Correct 'Peron(s)' to 'Person(s)'.",
      "exact_quote": "Sanctioned Peron(s)"
    },
    {
      "location": "Page 63, Article I (Definitions)",
      "error": "The phrase 'for the ratable benefit and the Secured Parties' contains a grammatical error. It should likely be 'for the ratable benefit of the Secured Parties'.",
      "suggestion": "Change 'benefit and the Secured Parties' to 'benefit of the Secured Parties'.",
      "exact_quote": "for the ratable benefit and the Secured Parties"
    },
    {
      "location": "Page 65, Article I (Definitions)",
      "error": "There is a pagination error. The page number is '29', but it should sequentially be '30' after the preceding page (Page 63) which was '28'.",
      "suggestion": "Correct the page number to '30'.",
      "exact_quote": "29"
    },
    {
      "location": "Page 65, Article I (Definitions)",
      "error": "The term 'Base Rate SOFR Determination Day' is used, but the defined term is 'Base Rate Term SOFR Determination Day'. This is an inconsistency or typo.",
      "suggestion": "Replace 'Base Rate SOFR Determination Day' with 'Base Rate Term SOFR Determination Day'.",
      "exact_quote": "Base Rate SOFR Determination Day"
    },
    {
      "location": "Page 66, Article I (Definitions)",
      "error": "There is a pagination error. The page number is '29', but it should sequentially be '31' after the preceding page (Page 65) which was '29'.",
      "suggestion": "Correct the page number to '31'.",
      "exact_quote": "29"
    },
    {
      "location": "Page 67, Article I (Definitions)",
      "error": "Section 4.2(a) is referenced in the definition of 'U.S. Government Securities Business Day', but Article IV is explicitly 'RESERVED' in this document and does not contain a Section 4.2(a). This is likely an incorrect cross-reference.",
      "suggestion": "Correct the cross-reference to the appropriate section, e.g., Section 6.2(a).",
      "exact_quote": "Sections 2.3(a), 2.4(c), 4.2(a) and 5.2"
    },
    {
      "location": "Page 75, Section 2.3(a)",
      "error": "The text uses '(1) U.S. Government Securities Business Day' where a numerical word 'one' is expected, consistent with other numerical spellings in parentheses (e.g., 'three (3)').",
      "suggestion": "Change '(1)' to 'one'.",
      "exact_quote": "no later than (1) U.S. Government Securities Business Day"
    },
    {
      "location": "Page 77, Section 2.4(a) and (b)",
      "error": "Paragraph (b) 'Mandatory Prepayments' is incorrectly formatted as a continuation of the sentence from paragraph (a) 'Repayment on Termination Date' instead of as a new, distinct sub-section.",
      "suggestion": "Start ' (b) Mandatory Prepayments.' on a new line and ensure it is properly formatted as a separate sub-section.",
      "exact_quote": "accrued but unpaid interest thereon (b) Mandatory Prepayments."
    },
    {
      "location": "Page 77, Section 2.4(d)",
      "error": "The section is marked 'Reserved', indicating a placeholder or an incomplete section.",
      "suggestion": "Either provide content for this section or explicitly state that it is intentionally left blank if it is not a drafting oversight.",
      "exact_quote": "(d) Reserved."
    },
    {
      "location": "Page 92, Section 5.9",
      "error": "Section 4.4(a) is referenced in the indemnity clause, but Article IV is explicitly 'RESERVED' in this document and does not contain a Section 4.4(a). This is likely an incorrect cross-reference.",
      "suggestion": "Correct the cross-reference to the appropriate section, e.g., Section 2.4(c).",
      "exact_quote": "Section 4.4(a)"
    }
  ]
}

def _get_gemini_client():
    """Returns a Gemini Client using API keys from environment variables."""
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        api_key = os.getenv("GOOGLE_API_KEY")
    
    if not api_key:
         logger.error("API Key Missing: GEMINI_API_KEY or GOOGLE_API_KEY must be set.")
         raise RuntimeError("GEMINI_API_KEY not set in environment variables.")
    
    return genai.Client(api_key=api_key)

def _get_openai_client():
    """Returns an OpenAI Client using API keys from environment variables."""
    api_key = os.getenv("OPENAI_API_KEY")
    
    if not api_key:
         logger.error("API Key Missing: OPENAI_API_KEY must be set.")
         raise RuntimeError("OPENAI_API_KEY not set in environment variables.")
    
    return openai.OpenAI(api_key=api_key)

# The prompt defines the AI's persona and strict output requirements.
# It uses a few-shot style instruction to ensure valid JSON format.
TEST_PROMPT = """
Role: You are a strict, pedantic Legal Proofreader. You are reviewing a standalone legal document fragment.
Input: The attached text from a contract.

CRITICAL INSTRUCTIONS:
1. **Assume Isolation with Common Sense**: Do NOT assume missing definitions exist in other documents. However, IGNORE common commercial lending terms typically defined in a base Credit Agreement (e.g., "Borrower", "Administrative Agent", "Lender", "Business Day", "Dollars", "GAAP", "Material Adverse Effect"). Only flag specific, deal-specific, or unusual capitalized terms that are undefined.
2. **Logic Check:** Check all math and logic tables.
3. **Drafting Errors:** Find any placeholders like "[__]" or blank lines that were forgotten.
4. **Exact Quotes:** For every error, you MUST provide the `exact_quote` from the text that contains the error. This is used for highlighting.

Output Format:
Return ONLY a valid JSON object with the following structure:
{
  "errors": [
    {
      "location": "Page 3, Section 2.1",
      "error": "Description of the error",
      "suggestion": "Suggested fix",
      "exact_quote": "Exact text substring to highlight in red"
    }
  ]
}
If no errors are found, return {"errors": []}.
"""

def _yield_log(level: str, message: str):
    """Utility to yield log objects in a format the frontend expects."""
    from datetime import datetime
    return json.dumps({
        "log": {
            "timestamp": datetime.now().strftime("%H:%M:%S.%f")[:-3],
            "level": level,
            "message": f"[{level}] {message}"
        }
    }) + "\n"

def _log_to_excel(model: str, filename: str, prompt: str, output: str):
    """Logs the analysis session to an Excel file."""
    file_path = "audit_logs.xlsx"
    headers = ["Timestamp", "Model", "Filename", "Prompt Snippet", "Full Prompt", "Output"]
    
    # Truncate prompt for easier viewing in snippet column
    prompt_snippet = (prompt[:100] + "...") if len(prompt) > 100 else prompt

    try:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            # Check if headers match, if not (old file), maybe append column? 
            # ideally we would check, but for now let's just append row. 
            # If columns mismatch, it might look weird.
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            model,
            filename,
            prompt_snippet,
            prompt,
            output
        ])
        wb.save(file_path)
        logger.info(f"Audit log saved to {file_path}")
    except Exception as e:
        logger.error(f"Failed to save audit log to Excel: {e}")


async def analyze_document_generator(file_path: str, test_mode: bool = False, model: str = DEFAULT_GEMINI_MODEL):
    """
    Async Generator that analyzes a PDF and yields status updates and logs.
    """
    filename = os.path.basename(file_path)
    logger.info(f"Starting analysis for: {filename} (Test Mode: {test_mode}, Model: {model})")

    yield _yield_log("INFO", f"Initializing analysis pipeline for {filename}")
    yield json.dumps({"stage": "extracting", "message": f"Extracting text from {filename}..."}) + "\n"

    if test_mode:
        yield _yield_log("INFO", "Test Mode is enabled. Intercepting API calls.")
        
        yield _yield_log("DEBUG", "Simulating PDF text extraction...")
        await asyncio.sleep(0.5) 
        yield _yield_log("INFO", "Text extraction complete. Character count: 1240")
        
        yield json.dumps({"stage": "distributing", "message": "Topic Distributor: Routing to Legal Reviewer Agent..."}) + "\n"
        yield _yield_log("INFO", "Topic Distributor selected: legal_reviewer_v1")
        
        await asyncio.sleep(1) 
        yield json.dumps({"stage": "analyzing", "message": f"Legal Reviewer: Auditing contract logic with {model}..."}) + "\n"
        yield _yield_log("INFO", f"Legal Reviewer sent content to {model}...")
        
        await asyncio.sleep(1) 
        yield _yield_log("DEBUG", f"{model} returned raw JSON. Validating schema...")
        yield json.dumps({"stage": "finalizing", "message": "Finalizing report..."}) + "\n"
        
        await asyncio.sleep(0.5)
        
        await asyncio.sleep(0.5)
        yield _yield_log("INFO", "Analysis successfully completed.")
        
        # Run locator on mock data too if file exists (or skip if dummy)
        # For actual file uploads in test mode, we can still run locator
        mock_data = MOCK_ANALYSIS_RESULT.copy()
        
        # Only try to locate if it's a real file we can read, otherwise mock rects?
        # Since test mode often uses a random PDF, locator might fail to find "Effective Date".
        # Let's try it anyway so "Test Mode" + "Real PDF" works.
        yield json.dumps({"stage": "locating", "message": "Locator Swarm: Pinpointing exact text locations..."}) + "\n"
        yield _yield_log("INFO", "Running Locator Agent on Mock Data...")
        try:
            import re
            for err in mock_data["errors"]:
                 location = err.get("location", "")
                 m = re.search(r"Page\s+(\d+)", location, re.IGNORECASE)
                 if m:
                     page_num = int(m.group(1))
                     snippet = err.get("exact_quote", "")
                     rects = find_text_coordinates(file_path, page_num, snippet)
                     if rects:
                         err["boundingBoxes"] = rects
        except Exception:
            pass # consistency with non-test mode

        yield json.dumps({"result": mock_data}) + "\n"
        return

    try:
        is_gemini = "gemini" in model.lower()
        is_openai = "gpt" in model.lower() or "o1" in model.lower() or "o3" in model.lower()
        
        client: Any = None
        if is_gemini:
             yield _yield_log("INFO", f"Initializing Gemini client for {model}...")
             client = _get_gemini_client()
        elif is_openai:
             yield _yield_log("INFO", f"Initializing OpenAI client for {model}...")
             client = _get_openai_client()
        else:
             # Fallback or error
             yield _yield_log("WARNING", f"Unknown model '{model}'. Defaulting to Gemini.")
             model = DEFAULT_GEMINI_MODEL
             is_gemini = True
             client = _get_gemini_client()
        
        try:
            yield _yield_log("DEBUG", f"Opening file stream: {file_path}")
            with open(file_path, "rb") as f:
                try:
                    reader = PdfReader(f)
                    text = ""
                    page_count = len(reader.pages)
                    yield _yield_log("INFO", f"PDF loaded. Total pages discovered: {page_count}")
                    
                    for i, page in enumerate(reader.pages):
                        extracted = page.extract_text() or ""
                        # Insert a clear page marker so the AI handles cross-page text correctly
                        text += f"\n\n--- [START OF PAGE {i+1}] ---\n"
                        text += extracted
                        yield _yield_log("DEBUG", f"Page {i+1} processed. ({len(extracted)} chars)")
                except Exception as pdf_err:
                     yield _yield_log("ERROR", f"PDF Read Error: {str(pdf_err)}")
                     yield json.dumps({"result": {"errors": [{"location": "Document", "error": f"Corrupt or unreadable PDF: {str(pdf_err)}", "suggestion": "Try repairing the PDF or export it again."}]}}) + "\n"
                     return
        
            if not text.strip():
                 yield _yield_log("ERROR", "Extraction failed. PDF text layer is empty.")
                 yield json.dumps({"result": {"errors": [{"location": "Document", "error": "Could not extract text from PDF.", "suggestion": "Ensure PDF is text-based, not scanned image."}]}}) + "\n"
                 return
        except Exception as e:
            # Catch file access errors (though less likely with tempfile)
            yield _yield_log("ERROR", f"File Access Error: {str(e)}")
            raise e

        yield _yield_log("INFO", f"Extraction successful. Total content: {len(text)} characters.")
        
        yield json.dumps({"stage": "distributing", "message": "Topic Distributor: Parsing sections and routing tasks..."}) + "\n"
        yield _yield_log("INFO", "Analyzing contract metadata and structure...")
        
        yield json.dumps({"stage": "analyzing", "message": f"Legal Reviewer: Critiquing contract clauses with {model}..."}) + "\n"
        yield _yield_log("INFO", f"Sending context window of {len(text)} tokens to {model} API...")
        
        # Run sync API call in threadpool to avoid blocking event loop
        loop = asyncio.get_running_loop()
        
        full_prompt = f"{TEST_PROMPT}\n\n--- CONTRACT TEXT BEGINS ---\n{text}\n--- CONTRACT TEXT ENDS ---"
        raw_output = ""
        
        if is_gemini:
            response = await loop.run_in_executor(None, lambda: client.models.generate_content(
                model=model,
                contents=full_prompt,
                config={"response_mime_type": "application/json"}
            ))
            raw_output = response.text
        elif is_openai:
            response = await loop.run_in_executor(None, lambda: client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "You are a helpful legal assistant. Output valid JSON only."}, # System prompt slightly redundant but safe
                    {"role": "user", "content": full_prompt}
                ],
                response_format={"type": "json_object"}
            ))
            raw_output = response.choices[0].message.content

        yield _yield_log("INFO", f"Analysis received from {model}.")
        
        # Log to Excel (Non-test mode only)
        _log_to_excel(model, filename, full_prompt, raw_output)
        
        yield _yield_log("DEBUG", f"Raw AI Output snippet: {raw_output[:100]}...")

        yield json.dumps({"stage": "finalizing", "message": "Synthesizing findings into structured report..."}) + "\n"

        try:
            yield _yield_log("DEBUG", "Parsing and normalizing JSON schema...")
            data = json.loads(raw_output)
            
            if isinstance(data, list):
                yield _yield_log("WARNING", "AI returned list without wrapper. Normalizing...")
                data = {"errors": data}
            
            # --- LOCATOR STEP (PARALLEL SWARM) ---
            yield json.dumps({"stage": "locating", "message": "Locator Swarm: Pinpointing exact text locations..."}) + "\n"
            yield _yield_log("INFO", "Dispatching Locator Swarm (Parallel Neighbor Search)...")
            errors = data.get("errors", [])
            
            # Prepare tasks
            locator_tasks = []
            for err in errors:
                import re
                location = err.get("location", "")
                m = re.search(r"Page\s+(\d+)", location, re.IGNORECASE)
                if m:
                    page_num = int(m.group(1))
                    snippet = err.get("exact_quote", "")
                    if snippet and snippet != "[__]":
                         locator_tasks.append({
                             "page": page_num,
                             "text": snippet,
                             "error_ref": err # Keep reference to update later
                         })
            
            if locator_tasks:
                 from app.pdf_locator import batch_locate_text
                 yield _yield_log("DEBUG", f"Swarming {len(locator_tasks)} targets with thread pool...")
                 
                 results = batch_locate_text(file_path, locator_tasks)
                 
                 # Map results back
                 # Since we passed objects by reference (error_ref) in the task, and we can access original_task in result
                 start_time = datetime.now()
                 success_count = 0
                 
                 for res in results:
                     if res["found"]:
                         # Update the specific error object
                         task = res["original_task"]
                         err_obj = task["error_ref"]
                         err_obj["boundingBoxes"] = res["rects"]
                         
                         # If page was different (drift), maybe update location string?
                         # Optional: err_obj["location"] += f" (Found on Page {res['page']})"
                         success_count += 1
                     else:
                         # Log failure for debug
                         task = res['original_task']
                         # logger.warning(f"Could not locate '{task['text'][:10]}...'")

                 yield _yield_log("INFO", f"Locator Swarm finished. Resolved {success_count}/{len(locator_tasks)} precise locations.")

            yield json.dumps({"result": data}) + "\n"
            
        except json.JSONDecodeError as jde:
            yield _yield_log("ERROR", f"JSON Parse Failed: {str(jde)}")
            yield json.dumps({"result": {
                "errors": [{
                    "location": "System", 
                    "error": "AI returned invalid JSON structure.", 
                    "suggestion": "Check logs for raw output or retry analysis."
                }]
            }}) + "\n"

    except Exception as e:
        yield _yield_log("CRITICAL", f"Analysis pipeline crashed: {str(e)}")
        yield json.dumps({"result": {
            "errors": [{
                "location": "System",
                "error": f"Internal process failed: {str(e)}",
                "suggestion": "Check backend logs and API configuration."
            }]
        }}) + "\n"

async def analyze_document(file_path: str, test_mode: bool = False) -> Dict:
    """Wrapper for async generator (for non-streaming use cases)."""
    gen = analyze_document_generator(file_path, test_mode)
    last_res = {}
    async for item in gen:
        data = json.loads(item)
        if "result" in data:
            last_res = data["result"]
    return last_res
