from openpyxl import load_workbook
import json

def extract_target_log():
    try:
        wb = load_workbook('backend/audit_logs.xlsx')
        ws = wb.active
        # Assuming the last row is the one we want, or the one with many errors.
        # Let's iterate backwards and find the first one with > 15 errors
        
        rows = list(ws.iter_rows(values_only=True))
        headers = [cell.value for cell in ws[1]]
        
        for row in reversed(rows[1:]): # Skip header
             data = dict(zip(headers, row))
             output = data.get('Output', '')
             if output and output.startswith('{'):
                 try:
                     json_data = json.loads(output)
                     errors = json_data.get('errors', [])
                     if len(errors) >= 15:
                         print("FOUND_JSON_START")
                         print(output)
                         print("FOUND_JSON_END")
                         return
                 except:
                     continue
        
        print("No matching log found.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    extract_target_log()
